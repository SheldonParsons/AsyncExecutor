import os
from io import BytesIO
from typing import Union
import openpyxl
from openpyxl.workbook import Workbook as OpenpyxlWorkbook
import asyncio
import warnings

from core.task_object.generate_object import GlobalOption

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

from core.ast_file.core import AstFile


class AstWorkbook:
    """
    openpyxl.workbook.Workbook  转发、委托类
    """

    def __init__(self, workbook: OpenpyxlWorkbook):
        self._workbook = workbook

    def __getattr__(self, name: str):
        return getattr(self._workbook, name)

    def get_sheet(self, index: Union[int, str, None] = None):
        if index is None:
            return self._workbook.active
        elif isinstance(index, int):
            if index == -1:
                return self._workbook.active
            elif index > 0:
                try:
                    return self._workbook.worksheets[index]
                except IndexError:
                    raise RuntimeError(f"错误：下标 {index} 的工作表不存在")
            else:
                raise RuntimeError("Sheet 下标需要大于0")
        elif isinstance(index, str):
            try:
                return self._workbook[index]
            except KeyError:
                raise RuntimeError(f"错误：名为 '{index}' 的工作表不存在")
        else:
            raise RuntimeError(f"错误：工作表索引类型错误")

    def get_rows(self, sheet_index: Union[int, str, None] = None):
        sheet = self.get_sheet(sheet_index)
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append(row)
        return rows

    def get_row(self, row_index: int = 0, sheet_index: Union[int, str, None] = None):
        sheet = self.get_sheet(sheet_index)
        row_cells = sheet[row_index]
        cells = []
        for cell in row_cells:
            cells.append(cell.value)
        return cells

    def get_column(self, column_index: Union[int, str, None] = None, sheet_index: Union[int, str, None] = None):
        sheet = self.get_sheet(sheet_index)
        if isinstance(column_index, int):
            cols_iterator = sheet.iter_cols(min_col=column_index, max_col=column_index, values_only=True)
            cols = []
            for col in cols_iterator:
                cols.append(col)
            return cols
        elif isinstance(column_index, str):
            column_cells = sheet[column_index]
            cells = []
            for cell in column_cells:
                cells.append(cell.value)
            return cells
        else:
            column_cells = sheet[1]
            cells = []
            for cell in column_cells:
                cells.append(cell.value)
            return cells


class AsyncExcel:

    def __init__(self, global_options=None):
        self.ast_file: Union[AstFile, None] = None
        self.workbook: Union[AstWorkbook, OpenpyxlWorkbook, None] = None
        self.cover_file = False
        self.global_options = global_options

    async def async_load(self, file: Union[BytesIO, AstFile[BytesIO, GlobalOption], str], filename: str = None):
        if isinstance(file, AstFile):
            self.ast_file = file
        elif isinstance(file, BytesIO):
            if filename is None:
                raise RuntimeError(f"async_load 函数在接收一个文件对象时必须填写参数 filename")
            self.ast_file = AstFile(self.global_options).load(file, filename)
        elif isinstance(file, str):
            self.ast_file = await AstFile(self.global_options).load_from_system(file)
        self.workbook = AstWorkbook(openpyxl.load_workbook(self.ast_file.file))
        return self.workbook

    async def async_save(self, cover_file=False):
        self.cover_file = cover_file
        self.ast_file.cover_file = cover_file
        filepath, object_unique_key = self.ast_file.get_filepath(self.ast_file.file.name)
        directory = os.path.dirname(filepath)
        await asyncio.to_thread(os.makedirs, directory, exist_ok=True)
        self.workbook.save(filepath)
        return object_unique_key
